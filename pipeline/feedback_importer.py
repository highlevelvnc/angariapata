"""
Feedback Importer · Sprint 2026-05-11

Reads the operator-edited XLSX (where Susana filled the "Estado", "Notas"
and "Data Contacto" columns) and writes the results back to the Lead table:

  Lead.crm_stage         ← Estado (mapped via _STAGE_FROM_LABEL)
  Lead.contact_outcome   ← Notas
  Lead.last_contacted_at ← Data Contacto (DD/MM/YYYY parsed)
  Lead.contacted_by      ← --by CLI flag (default "Susana")
  Lead.re_engage_after   ← +30d if crm_stage == "sem_resposta"
                          ← cleared otherwise
  Lead.archived          ← True if "nao_interessado" or "convertido"

The join key is the "Lead ID" column (last column added by the export).
If a row doesn't have a numeric Lead ID, it's skipped silently.

CLI:
    python main.py import-feedback <path-to-edited.xlsx>
    python main.py import-feedback exports/comercial.xlsx --by "Susana"
    python main.py import-feedback path.xlsx --dry-run

Sheet handling: reads every sheet that has a "Lead ID" column. So both
"Lista Premium" and "Lista Expandida" get processed in a single call.
"""
from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from storage.database import get_db
from storage.models import Lead
from utils.logger import get_logger
from reports.commercial_export import _STAGE_FROM_LABEL

log = get_logger(__name__)


# Stages that trigger automatic archive (the lead is done either way)
_ARCHIVE_STAGES = {"nao_interessado", "convertido"}

# Stages that trigger re-engagement (re-surface in N days)
_RE_ENGAGE_DAYS = 30
_RE_ENGAGE_STAGES = {"sem_resposta", "contactado"}


def _parse_date(value: object) -> Optional[datetime]:
    """Accept datetime, ISO string, DD/MM/YYYY, or empty."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    log.debug("[feedback] couldn't parse date {v!r}", v=s)
    return None


def _map_stage(label: str) -> Optional[str]:
    """Convert XLSX label (PT) → canonical crm_stage value."""
    if not label:
        return None
    s = str(label).strip().lower().lstrip("✓").strip()
    return _STAGE_FROM_LABEL.get(s)


def import_feedback(xlsx_path: str, contacted_by: str = "Susana",
                    dry_run: bool = False) -> dict:
    """
    Read every sheet of the XLSX and update matching leads in the DB.

    Returns stats dict:
        {sheets_read, rows_seen, updated, archived, re_engagement_scheduled,
         unchanged, missing_lead_id, errors}
    """
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")

    wb = load_workbook(path, data_only=True)
    stats = {
        "sheets_read":              0,
        "rows_seen":                0,
        "updated":                  0,
        "archived":                 0,
        "re_engagement_scheduled":  0,
        "unchanged":                0,
        "missing_lead_id":          0,
        "missing_estado":           0,
        "lead_not_found":           0,
    }

    now = datetime.utcnow()

    with get_db() as db:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [c.value for c in ws[1]]
            if "Lead ID" not in headers and "lead_id" not in headers:
                continue
            stats["sheets_read"] += 1

            # Build column index (lower-cased keys for safety)
            col_idx = {(h or "").strip().lower(): i for i, h in enumerate(headers)}
            estado_col = col_idx.get("estado")
            notas_col  = col_idx.get("notas")
            data_col   = col_idx.get("data contacto") or col_idx.get("data_contacto")
            id_col     = col_idx.get("lead id") or col_idx.get("lead_id")

            if id_col is None:
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                stats["rows_seen"] += 1
                lead_id = row[id_col]
                try:
                    lead_id = int(lead_id) if lead_id else None
                except (ValueError, TypeError):
                    lead_id = None
                if not lead_id:
                    stats["missing_lead_id"] += 1
                    continue

                lead = db.query(Lead).filter(Lead.id == lead_id).first()
                if not lead:
                    stats["lead_not_found"] += 1
                    continue

                estado_raw = row[estado_col] if estado_col is not None else None
                notas_raw  = row[notas_col]  if notas_col  is not None else None
                data_raw   = row[data_col]   if data_col   is not None else None

                new_stage = _map_stage(estado_raw)
                # "Novo" is the default — skip if Susana didn't actually change anything
                if not new_stage or new_stage == "novo":
                    stats["missing_estado"] += 1
                    continue

                # Compute new values
                contact_date = _parse_date(data_raw) or now

                # Only update if something actually changed (idempotent)
                changed = False
                if lead.crm_stage != new_stage:
                    lead.crm_stage = new_stage
                    changed = True
                if notas_raw and lead.contact_outcome != str(notas_raw).strip():
                    lead.contact_outcome = str(notas_raw).strip()
                    changed = True
                if not lead.last_contacted_at or lead.last_contacted_at < contact_date:
                    lead.last_contacted_at = contact_date
                    changed = True
                if lead.contacted_by != contacted_by:
                    lead.contacted_by = contacted_by
                    changed = True

                # Side effects
                if new_stage in _ARCHIVE_STAGES:
                    if not lead.archived:
                        lead.archived = True
                        stats["archived"] += 1
                        changed = True
                    lead.re_engage_after = None
                elif new_stage in _RE_ENGAGE_STAGES:
                    lead.re_engage_after = contact_date + timedelta(days=_RE_ENGAGE_DAYS)
                    stats["re_engagement_scheduled"] += 1
                    changed = True
                else:
                    # interessado / indisponivel — no re-engagement schedule
                    lead.re_engage_after = None

                if changed:
                    stats["updated"] += 1
                else:
                    stats["unchanged"] += 1

        if dry_run:
            db.rollback()
            log.info("[feedback] dry-run — no DB writes")
        else:
            db.commit()

    log.info(
        "[feedback] read {sheets} sheets, {rows} rows · updated={u} · "
        "archived={a} · re-engagement +30d={r} · unchanged={x} · "
        "missing_id={mi} · missing_estado={me} · not_found={nf}",
        sheets=stats["sheets_read"], rows=stats["rows_seen"],
        u=stats["updated"], a=stats["archived"],
        r=stats["re_engagement_scheduled"], x=stats["unchanged"],
        mi=stats["missing_lead_id"], me=stats["missing_estado"],
        nf=stats["lead_not_found"],
    )
    return stats
