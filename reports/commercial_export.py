"""
Commercial Export — lista comercial pronta para entrega ao cliente.

Gera um único XLSX com três folhas:
  1. "Lista Premium"    — proprietários directos, mobile/relay, score alto
  2. "Lista Expandida"  — oportunidades adicionais, mais abrangente
  3. "Resumo Executivo" — KPIs prontos para apresentação ao cliente

Critérios Lista Premium (ordenada: mobile primeiro → relay → score desc):
  • telefone válido obrigatório (mobile ou relay preferido; sem fixos 21x/22x)
  • agências excluídas (owner_type="agency" ou lead_type="agency_listing")
  • tipos permitidos: fsbo, frbo, active_owner, unknown
  • score ≥ warm_threshold (configurável, default: settings.warm_score_threshold)
  • zonas-alvo do .env (configurável)
  • deduplicação por telefone (mantém score mais alto)
  • limite configurável (default: 50)

Critérios Lista Expandida:
  • telefone válido obrigatório
  • agências confirmadas excluídas
  • score ≥ (warm_threshold - 10), floor=25
  • exclui fixos Lisboa/Porto (21x/22x) → provável agência
  • remove telefones já presentes na Lista Premium
  • deduplicação por telefone
  • limite configurável (default: 150)

Uso:
  python main.py export-commercial
  python main.py export-commercial --premium-limit 30 --expanded-limit 100
  python main.py export-commercial --zones Lisboa,Cascais
  python main.py export-commercial --output-dir exports/
"""
from __future__ import annotations

import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Optional

from config.settings import settings
from storage.database import get_db
from storage.models import Lead
from utils.logger import get_logger

# Re-use helpers from contact_export to avoid duplication
from reports.contact_export import (
    _canonical_phone,
    _whatsapp_link,
    _get_url,
    _build_insight,
    _is_likely_agency,
    _format_price,
    _confidence_label,
    _LEAD_TYPE_LABELS,
    _SOURCE_LABELS,
)

log = get_logger(__name__)

# ── Phone type sort priority (mobile first) ───────────────────────────────────
_PHONE_PRIORITY = {"mobile": 0, "relay": 1, "unknown": 2, "landline": 3}

# ── Phone type labels ─────────────────────────────────────────────────────────
_PHONE_LABELS = {
    "mobile":   "Telemóvel",
    "relay":    "Relay/OLX",
    "landline": "Fixo",
    "unknown":  "Desconhecido",
}


# ── Owner Tier classification ──────────────────────────────────────────────────

def _phone_uses_count(phone: str) -> int:
    """Return how many other leads share this phone (cheap-cached probe)."""
    if not phone or not hasattr(_phone_uses_count, "_cache"):
        _phone_uses_count._cache = {}    # type: ignore
    cache = _phone_uses_count._cache    # type: ignore
    if phone in cache:
        return cache[phone]
    with get_db() as db:
        from sqlalchemy import func
        n = (
            db.query(func.count(Lead.id))
            .filter(Lead.contact_phone == phone, Lead.archived == False)
            .scalar() or 0
        )
    cache[phone] = n
    return n


# Substring markers — match anywhere in the lowercased name.
_CORP_MARKERS: tuple[str, ...] = (
    "lda", "s.a", "imobiliária", "imobiliaria", "investimentos",
    "properties", "property", "consultora", "consultor",
    "mediadora", "mediação", "mediacao", "vendedor profissional",
    "lisabona", "century", "remax", "re/max", "savills", "engel",
    "sotheby", "predimed", "sociedade", "unipessoal", "amplitude",
    "habivida", "knight", "real estate", "realestate", "patabrava",
    "construções", "construcoes", "investments", "holding", "group",
)
# Whole-token markers — match a token exactly (case-insensitive).
# These would generate false positives if matched as substrings (e.g. "kw"
# would match "Akwa", "iad" would match "ciada"). Matching the token alone
# catches "KW PRO", "IAD Portugal", "Casas ERA", etc. without false hits.
_CORP_TOKENS: frozenset[str] = frozenset({
    "kw", "iad", "era", "zome", "ami", "sa", "imo", "casas",
    "imoveis", "imóveis", "predial", "predior", "mediação",
    # Country/region tokens used as agency suffixes ("Homebook Portugal",
    # "Casas Portugal", "Group HBC"). Real PT people are not surnamed Portugal.
    "portugal", "brasil", "lisboa", "porto",
    # Generic agency suffixes
    "lda", "unipessoal", "pro", "team", "homes", "house",
})
_PERSONAL_NAME_RE = re.compile(r"^[A-Za-zÀ-ÿ' -]{2,40}$")

# Whitelist of common Portuguese first names (M+F). When the first token of a
# value is in this set, we treat the whole value as a personal name even if
# subsequent tokens are unfamiliar surnames. This is the strongest signal —
# stronger than corporate-marker detection which is necessarily a blacklist.
_PT_FIRST_NAMES: frozenset[str] = frozenset({
    # Top 200 PT first names (mix M/F, simplified)
    "ana","maria","joão","joao","jose","josé","manuel","antonio","antónio",
    "francisco","carlos","paulo","pedro","luis","luís","miguel","rui",
    "fernando","ricardo","jorge","rita","sofia","catarina","ines","inês",
    "marta","beatriz","mariana","leonor","margarida","filipa","raquel",
    "patricia","patrícia","sara","helena","cristina","susana","alexandra",
    "isabel","teresa","fatima","fátima","graca","graça","ricarda",
    "tiago","bruno","andre","andré","nuno","diogo","filipe","goncalo",
    "gonçalo","artur","afonso","duarte","henrique","hugo","ivan","joel",
    "leonardo","luciano","lucas","marco","marcos","martim","mauricio",
    "octavio","octávio","oliveira","raul","rodrigo","ruben","rúben",
    "samuel","sebastiao","sebastião","sergio","sérgio","simao","simão",
    "tomas","tomás","valter","vasco","vicente","vitor","vítor","xavier",
    "alberto","alfredo","alvaro","álvaro","angelo","ângelo","armando",
    "augusto","aurelio","aurélio","bernardo","cesar","césar","claudio",
    "cláudio","daniel","david","dinis","domingos","edgar","eduardo",
    "elisio","emanuel","ernesto","estevao","estêvão","fabio","fábio",
    "felisberto","fernao","fernão","gabriel","gaspar","geraldo","gilberto",
    "guilherme","ilidio","jaime","jacinto","joaquim","leandro","leonel",
    "loureiro","mario","mário","matias","mauro","mateus","nelson","nicolau",
    "norberto","octaviano","olegario","olegário","paulino","quintino",
    "raimundo","ramiro","reinaldo","renato","romeo","romeu","ronaldo",
    "rosario","rosário","salvador","santiago","saul","silvestre","silvio",
    "sílvio","teodoro","teófilo","timoteo","timóteo","tobias","ubaldo",
    "urbano","valdemar","valentim","veloso","virgilio","virgílio",
    "wilson","ze","zé","zeferino",
    # F
    "alice","aline","alina","amalia","amanda","andreia","angela","ângela",
    "antonia","antónia","aurora","barbara","bárbara","branca","brigida",
    "brígida","camila","carla","carlota","carmen","carolina","celeste",
    "celia","célia","cidalia","cidália","clara","clarisse","claudia",
    "cláudia","cristiana","daniela","debora","débora","diana","dora",
    "dulce","edite","elena","eliana","elsa","elvira","emilia","emília",
    "estela","ester","etelvina","eulalia","eulália","eunice","eva",
    "fabiola","fabíola","fernanda","filomena","flavia","flávia","floriana",
    "francesca","gabriela","gisela","glória","gloria","guilhermina",
    "iara","ida","idalina","ilda","ines","inês","irene","iris","íris",
    "joana","julia","júlia","juliana","julieta","laura","leticia",
    "letícia","lia","liana","liliana","linda","lucia","lúcia","luciana",
    "luisa","luísa","luiza","madalena","manuela","matilde","melania",
    "melânia","melanie","melissa","michelle","milena","miriam","mónica",
    "monica","natacha","natalia","natália","nina","noemia","noémia",
    "nora","odete","olga","palmira","paula","pilar","preciosa","raquel",
    "renata","romana","rosa","rosalia","rosália","rosana","rosario",
    "rute","sandra","silvana","silvia","sílvia","sonia","sónia","stela",
    "tamara","tania","tânia","telma","tina","valentina","vanda","vanessa",
    "vera","veronica","verónica","violeta","virgínia","virginia","vitoria",
    "vitória","yara","zélia","zelia","zita",
    # Brazilian/lusophone variants common in PT-PT
    "alex","camelia","camélia","alinne","camilla","fabricio","fabrício",
    "jorge","kelly","wagner","adilson","cleber","reginaldo","valdir",
    "rosanilopes","catarinagaio",  # leaked usernames already accepted by their first part
})

# Portal UI noise that leaks into contact_name fields when the scraper picks
# up button labels or section headers instead of an actual seller name.
_NAME_NOISE: frozenset[str] = frozenset({
    "adicionados hoje", "adicionado hoje", "adicionados ontem",
    "vendedor profissional", "vendedor", "particular", "particulares",
    "anuncio", "anúncio", "anuncios", "anúncios", "ver telefone",
    "ver número", "contactar", "novo", "nova", "ontem", "hoje",
    "destaque", "destaques", "premium",
})


def _is_personal_name(value: str) -> bool:
    """
    Return True if `value` looks like a personal name (e.g. "Maria",
    "Ana Sousa", "Artur Santos") rather than a corporate / agency name.

    Used to demote false-positive Tier C leads where the seller's first
    name was extracted into agency_name during scraping.
    """
    v = (value or "").strip()
    if not v:
        return False
    low = v.lower()
    # Reject portal UI noise like "Adicionados hoje", "Vendedor", etc.
    if low in _NAME_NOISE:
        return False
    # Fast-reject digits, parens, slashes, ampersand — agencies / flipper tags
    if any(ch.isdigit() for ch in v):
        return False
    if any(ch in v for ch in "()[]/&@|"):
        return False
    if any(m in low for m in _CORP_MARKERS):
        return False
    # Personal names: 1-3 tokens, only letters/spaces/'/-, total ≤ 40 chars
    tokens = v.split()
    if not (1 <= len(tokens) <= 3):
        return False
    # Reject if ANY token is a known corporate brand/keyword
    if any(t.lower().rstrip(".,") in _CORP_TOKENS for t in tokens):
        return False
    if not _PERSONAL_NAME_RE.match(v):
        return False
    # Single token: REQUIRES dict match (no surname to disambiguate).
    # Two/three tokens: dict match OR all tokens proper-cased lowercase letters
    # (covers "Cristina Sousa", "Eddy Costa" — names we don't have in our dict
    # but pattern-match as personal). The corporate-marker filter above already
    # rejected most agency strings.
    first = tokens[0].lower().rstrip(".,'-")
    if first in _PT_FIRST_NAMES:
        return True
    if len(tokens) == 1:
        return False
    # Multi-token fallback: every token starts uppercase and the rest is
    # lowercase letters. Rejects "KW PRO", "IAD Portugal", "Group HBC"
    # (have all-caps tokens that aren't title-cased), "Oferta privada"
    # (second token lowercase). Accepts "Cristina Sousa", "Eddy Costa".
    for t in tokens:
        if not t[0].isupper():
            return False
        if len(t) > 1 and not t[1:].islower():
            return False
    return True


def _effective_agency_name(lead: Lead) -> str:
    """
    Return agency_name only if it really looks corporate. Personal names
    leaked into the field (e.g. "Maria", "Marco", "Ana Sousa") return "".
    """
    raw = (lead.agency_name or "").strip()
    if not raw:
        return ""
    if _is_personal_name(raw):
        return ""
    return raw


def _classify_owner_tier(lead: Lead) -> str:
    """
    Classify how confident we are that this phone is THE OWNER.

      ✅ A · mobile + FSBO + único + nome    → quase certeza dono
      🟢 B · mobile + FSBO + 1-2 usos        → provável dono
      🟡 C · mobile com agency_name          → intermediário comercial
      🟡 D · landline + FSBO sem agency_name → casa do dono ou senhorio (50/50)
      ❌ E · landline com agency_name OU 4+ usos → switchboard agência
      ⚠ ?  · qualquer outro caso
    """
    pt   = (lead.phone_type or "").lower()
    is_o = bool(lead.is_owner)
    ag   = _effective_agency_name(lead)
    nm   = (lead.contact_name or "").strip()
    # If agency_name was actually a personal name, treat lead as FSBO-leaning
    # for the purposes of this classifier — fixes the Tier C false-positives
    # where scrapers leaked first names ("Maria", "Marco") into agency_name.
    if not ag and (lead.agency_name or "").strip():
        is_o = True
        ot_override = "fsbo"
    else:
        ot_override = None
    ot   = ot_override or (lead.owner_type or "").lower()
    n_uses = _phone_uses_count(lead.contact_phone or "")

    # Hard reject: landline in 4+ listings is switchboard
    if pt == "landline" and n_uses >= 4:
        return "E"

    if pt == "mobile" and is_o and ot == "fsbo" and not ag and n_uses == 1 and nm and len(nm) >= 3:
        return "A"
    if pt == "mobile" and is_o and ot == "fsbo" and not ag and n_uses <= 2:
        return "B"
    if pt == "mobile" and ag:
        return "C"
    # Mobile flagged as agency without a corp name — treat as flipper/agent
    # disguised. ≥3 reuses = switchboard E, ≤2 = weak agency C.
    if pt == "mobile" and ot == "agency" and not ag:
        return "E" if n_uses >= 3 else "C"
    # Mobile owner with unknown owner_type but ≤3 uses → likely owner (B)
    if pt == "mobile" and is_o and ot in ("unknown", "") and n_uses <= 3:
        return "B"
    # Relay numbers (OLX-style) follow the same logic as mobile but capped
    if pt == "relay" and is_o and ot == "fsbo" and not ag and n_uses <= 2:
        return "B"
    if pt == "relay" and (ag or n_uses >= 4):
        return "E"
    if pt == "relay":
        return "D"
    if pt == "landline" and not ag and n_uses <= 2:
        return "D"
    if pt == "landline" and (ag or n_uses >= 3):
        return "E"
    return "?"


# ── Personalised WhatsApp message ──────────────────────────────────────────────
def _first_name(name: str | None) -> str:
    if not name:
        return ""
    parts = name.strip().split()
    return parts[0].title() if parts else ""


def _short_zone(zone: str | None) -> str:
    """Return a clean zone name without parish-cascade noise."""
    if not zone:
        return "Lisboa"
    z = zone.strip()
    # Strip everything after first comma (parish cascade like "Estrela, Lisboa, Lisboa")
    if "," in z:
        z = z.split(",")[0].strip()
    # Strip "Lisboa-" prefix
    z = z.replace("Lisboa-", "").replace("-", " ").strip()
    return z or "Lisboa"


def _personalised_whatsapp(lead: Lead) -> tuple[str, str]:
    """
    Build (message, wa_link) personalised by typology, zone and lead_type.

    Tone: friendly, direct, PT-PT, name the lead by context.
    Designed for the Patabrava agency (Lisbon luxury). Keep under 320 chars
    so it fits WhatsApp preview without truncation.
    """
    from urllib.parse import quote

    phone   = _canonical_phone(lead.contact_phone or "") or (lead.contact_phone or "")
    if not phone:
        return ("", "")

    fn   = _first_name(lead.contact_name)
    typ  = (lead.typology or "").strip()
    pt   = (lead.property_type or "").strip().lower()
    zone = _short_zone(lead.zone)
    lt   = (lead.lead_type or "").lower()
    price = lead.price

    # Subject reference
    if pt == "terreno" or typ == "Terreno":
        subj = f"o seu terreno em {zone}"
    elif pt == "moradia" or "moradia" in (lead.title or "").lower():
        subj = f"a sua moradia em {zone}"
    elif typ:
        subj = f"o seu {typ} em {zone}"
    else:
        subj = f"o seu imóvel em {zone}"

    # Action verb
    if lt == "frbo":
        action = "arrendar"
    else:
        action = "vender"

    greet = f"Olá {fn}!" if fn else "Boa tarde!"

    msg = (
        f"{greet} Vi o seu anúncio sobre {subj} e fiquei interessado(a). "
        f"Sou da Pata Brava — trabalhamos exactamente nesta zona com clientes "
        f"que procuram este perfil. Está aberto(a) a uma breve conversa "
        f"sobre {action}? Obrigado!"
    )

    digits = phone.replace("+", "").replace(" ", "")
    wa_link = f"https://wa.me/{digits}?text={quote(msg)}"
    return (msg, wa_link)


# ── Motivation badges ──────────────────────────────────────────────────────────
def _motivation_badges(lead: Lead) -> list[str]:
    """
    Detect signals that hint at a motivated seller. Returns a list of
    short emoji-prefixed labels suitable for an XLSX column or HTML chip.

    Signals (active today; days_on_market + price_drop activate after 2nd run):
      🔗 EM 2+ PORTAIS — listed in OLX and Imovirtual simultaneously
      ⭐ ELITE         — score 90+ (top 5 % of HOT)
      👤 PROPRIETÁRIO  — confirmed FSBO (no agency)
      💎 LUXURY        — price ≥ 500 000 € (Patabrava sweet spot)
      📍 ZONA PRIME    — Estrela / Misericórdia / Santo Antonio / Cascais centro
      🔥 BAIXOU PREÇO  — price drop >5 % vs 30 d ago (only after run #2)
      📅 LONGO MERCADO — days_on_market > 60 (only after run #2)
    """
    out: list[str] = []

    # 🔗 Multi-portal
    sj = (lead.sources_json or "")
    has_imo = "imovirtual" in sj.lower()
    has_olx = '"olx"' in sj.lower() or "olx.pt" in sj.lower()
    if has_imo and has_olx:
        out.append("🔗 EM 2 PORTAIS")

    # ⭐ Elite score
    if (lead.score or 0) >= 90:
        out.append("⭐ ELITE")

    # 👤 FSBO
    if getattr(lead, "is_owner", False) and (lead.lead_type or "").lower() in ("fsbo","frbo","active_owner"):
        if not lead.agency_name:
            out.append("👤 PROPRIETÁRIO")

    # 💎 Luxury price (sale only)
    if lead.price and lead.price >= 500_000 and (lead.lead_type or "").lower() != "frbo":
        out.append("💎 LUXURY")

    # 📍 Prime zone
    z = (lead.zone or "").lower()
    parish = (lead.parish or "").lower()
    prime_keywords = (
        "estrela","misericórdia","misericordia","santo antónio","santo antonio",
        "lapa","chiado","príncipe real","principe real","baixa","alfama",
        "cascais centro","estoril","quinta da marinha","monte estoril",
    )
    if any(k in z or k in parish for k in prime_keywords):
        out.append("📍 ZONA PRIME")

    # 🔥 Price drop (requires history)
    if lead.price_changes and lead.price_changes != "{}":
        out.append("🔥 PREÇO BAIXOU")

    # 📅 Long on market (only meaningful after multiple runs)
    if (lead.days_on_market or 0) >= 60:
        out.append("📅 LONGO MERCADO")

    # 🔄 Re-listed (Sprint Engine C) — strong motivation signal
    rl = getattr(lead, "re_list_count", 0) or 0
    if rl >= 1:
        out.append(f"🔄 RE-LISTADO {rl}×")

    # ✓ Confidence tier (Sprint Integrity N)
    conf = getattr(lead, "contact_confidence", 0) or 0
    if conf >= 80:
        out.append(f"✓ ALTA CONFIANÇA {conf}")
    elif conf >= 60:
        out.append(f"◔ MÉDIA CONFIANÇA {conf}")
    # baixa confiança: NÃO se exibe (silêncio = aviso)

    # ⏰ Urgency + reason (Sprint NLP X)
    try:
        import json as _json
        breakdown = _json.loads(lead.score_breakdown or "{}")
    except Exception:
        breakdown = {}
    urg = breakdown.get("urgency", 0)
    if urg >= 6:
        out.append(f"⏰ URGÊNCIA {urg}/10")
    reason_emoji = {
        "EMIGRAÇÃO":   "🏃",
        "DIVÓRCIO":    "💔",
        "HERANÇA":     "📜",
        "EXECUÇÃO":    "⚖️",
        "REMODELAÇÃO": "🔨",
        "INVESTIDOR":  "📈",
        "INVESTIMENTO":"💼",
    }
    reason = breakdown.get("reason")
    if reason and reason in reason_emoji:
        out.append(f"{reason_emoji[reason]} {reason}")

    # 🆕 Recency (Sprint Engine HH) · listings frescos convertem mais
    if lead.first_seen_at:
        from datetime import datetime as _dt
        age_h = (_dt.utcnow() - lead.first_seen_at).total_seconds() / 3600
        if age_h <= 24:
            out.append("🆕 NOVO HOJE")
        elif age_h <= 168:  # 7 days
            out.append("🌟 ESTA SEMANA")

    # 👥 Multi-property owner (Sprint Engine WW)
    try:
        from pipeline.owner_profile import get_profile
        p = get_profile(lead)
        if p.is_multi_property and p.listings_count >= 2:
            out.append(f"👥 PORTFÓLIO {p.listings_count}×")
    except Exception:
        pass

    return out

# ── Commercial insight — richer than generic insight ─────────────────────────
_URGENCY_RE: list[tuple[re.Pattern, str]] = [
    (re.compile(r"\burgente\b|\burgência\b", re.I),              "venda urgente"),
    (re.compile(r"\bherança\b|\bherdeiro\b|\bpartilha\b", re.I), "herança/partilha"),
    (re.compile(r"\bdivórcio\b|\bseparação\b", re.I),            "divórcio"),
    (re.compile(r"\bemigr\w+\b", re.I),                          "proprietário a emigrar"),
    (re.compile(r"\bpenhora\b|\bexecutiv\w+\b", re.I),           "execução hipotecária"),
    (re.compile(r"\bpermuta\b", re.I),                           "aberto a permuta"),
    (re.compile(r"\bnegocia\w*\b", re.I),                        "preço negociável"),
    (re.compile(r"sem comissão|direto do dono|proprietário vende", re.I),
                                                                  "direto do proprietário"),
    (re.compile(r"\bpreciso vender\b|\bprecisa vender\b", re.I), "precisa vender"),
    (re.compile(r"\bpara recuperar\b|\bpara remodelar\b", re.I), "imóvel para remodelar"),
]


def _commercial_insight(lead: Lead) -> str:
    """
    Build a single-line commercial insight for the client list.

    Format: [tipo_proprietário] · [canal] · [sinal principal]

    Examples:
      "Proprietário Venda · Telemóvel · 23% abaixo do mercado"
      "Proprietário Arrendamento · Telemóvel · senhorio directo"
      "Potencial Proprietário · Relay/OLX · 97 dias no mercado"
    """
    parts: list[str] = []

    # 1. Lead type label
    lt = lead.lead_type or "unknown"
    lbl = _LEAD_TYPE_LABELS.get(lt, lt)
    parts.append(lbl)

    # 2. Phone channel
    pt = getattr(lead, "phone_type", None) or "unknown"
    parts.append(_PHONE_LABELS.get(pt, "Desconhecido"))

    # 3. Primary signal (first match wins)
    signal = ""

    # Price delta takes priority
    delta = lead.price_delta_pct
    if delta and delta >= 15:
        signal = f"{delta:.0f}% abaixo do mercado"
    elif delta and delta >= 5:
        signal = f"{delta:.0f}% abaixo do benchmark"

    # Urgency keywords
    if not signal:
        text = f"{lead.title or ''} {lead.description or ''}"
        for pat, label in _URGENCY_RE:
            if pat.search(text):
                signal = label
                break

    # Days on market
    if not signal:
        dom = lead.days_on_market or 0
        if dom >= 90:
            signal = f"{dom} dias no mercado"
        elif dom >= 45:
            signal = f"{dom} dias listado"

    # FRBO specific
    if not signal and lt == "frbo":
        signal = "senhorio a arrendar directamente"

    # Fallback — mobile confirmation
    if not signal and pt == "mobile":
        signal = "telemóvel directo confirmado"

    if signal:
        parts.append(signal)

    return " · ".join(parts)


def _is_excluded_landline(lead: Lead) -> bool:
    """Return True for Lisboa/Porto landlines — almost always agency lines."""
    phone = (lead.contact_phone or "").replace("+351", "").strip()
    return phone[:2] in ("21", "22")


# ── List builders ─────────────────────────────────────────────────────────────

def _lead_to_row(lead: Lead, rank: int | None = None) -> dict:
    """Convert a Lead ORM object to an export row dict."""
    phone     = _canonical_phone(lead.contact_phone or "") or lead.contact_phone or ""
    pt        = getattr(lead, "phone_type", None) or "unknown"
    tipo_tel  = _PHONE_LABELS.get(pt, "Desconhecido")
    url       = _get_url(lead.sources_json or "[]") or (lead.url if hasattr(lead, "url") else "")
    conf      = _confidence_label(lead)
    insight   = _commercial_insight(lead)
    wa_msg, wa_link = _personalised_whatsapp(lead)
    badges    = _motivation_badges(lead)

    # Sprint Engine WW · owner profile summary
    try:
        from pipeline.owner_profile import get_profile, format_profile_summary
        owner_summary = format_profile_summary(get_profile(lead))
    except Exception:
        owner_summary = "—"

    # Sprint Owner Tier · classify how confident we are this is the actual owner
    owner_tier = _classify_owner_tier(lead)

    # Sprint Pre-call Briefing · only built for Tier A/B (cheap to skip the rest)
    if owner_tier in ("A", "B"):
        try:
            from pipeline.precall_briefing import build_briefing
            _br = build_briefing(lead)
            briefing_text = _br["text"]
            opening_line  = _br["opening"]
            commission_eur = _br["commission_eur"]
        except Exception:
            briefing_text = ""
            opening_line  = ""
            commission_eur = 0
    else:
        briefing_text = ""
        opening_line  = ""
        commission_eur = 0

    row = {
        "rank":          rank,
        "score":         lead.score or 0,
        "label":         lead.score_label or "COLD",
        "owner_tier":    owner_tier,
        "badges":        " · ".join(badges) if badges else "—",
        "owner_profile": owner_summary,
        "nome":          lead.contact_name or "—",
        "telefone":      phone,
        "tipo_telefone": tipo_tel,
        "whatsapp":      wa_link,
        "mensagem_wa":   wa_msg,
        "zona":          lead.zone or "—",
        "concelho":      lead.municipality or lead.zone or "—",
        "tipologia":     lead.typology or "—",
        "preco":         _format_price(lead.price),
        "area_m2":       f"{lead.area_m2:.0f} m²" if lead.area_m2 else "—",
        "tipo_lead":     _LEAD_TYPE_LABELS.get(lead.lead_type or "unknown", lead.lead_type or "—"),
        "fonte":         _SOURCE_LABELS.get(lead.discovery_source or "", lead.discovery_source or "—"),
        "confianca":     conf,
        "insight":       insight,
        "url":           url,
        "dias_mercado":  lead.days_on_market or 0,
        "data_captacao": lead.first_seen_at.strftime("%d/%m/%Y") if lead.first_seen_at else "—",
        "briefing":      briefing_text or "—",
        "opening":       opening_line or "—",
        "comissao_est":  f"€{commission_eur:,.0f}".replace(",", " ") if commission_eur else "—",
        # internal fields for filtering/sorting — not written to sheet
        "_phone_type":   pt,
        "_lead_id":      lead.id,
        "_owner_type":   lead.owner_type or "",
    }
    return row


def generate_premium_list(
    score_min: int | None = None,
    zones: list[str] | None = None,
    limit: int = 50,
    min_confidence_score: int = 0,
) -> list[dict]:
    """
    Build the Lista Premium.

    Strict criteria:
      • score ≥ max(score_min, warm_threshold)
      • owner confirmed or unknown — agencies excluded
      • phone required, valid, not a Lisboa/Porto agency landline
      • FSBO/FRBO/active_owner/unknown lead types
      • target zones only (from .env unless overridden)
      • sorted: mobile first → relay → score desc
      • dedup by phone

    Returns list of row dicts, up to `limit` entries.
    """
    from sqlalchemy import select, and_, or_

    warm_t    = settings.warm_score_threshold
    min_score = max(score_min if score_min is not None else warm_t, warm_t)
    target_zones = zones or settings.zones

    with get_db() as db:
        q = (
            select(Lead)
            .where(
                Lead.archived == False,                       # noqa: E712
                Lead.is_demo  == False,                       # noqa: E712
                Lead.contact_phone.isnot(None),
                Lead.contact_phone != "",
                Lead.score >= min_score,
                # Sprint Quality C — exclude suspicious
                (Lead.listing_status.is_(None)) | (Lead.listing_status != "suspicious"),
                # Sprint REAL OWNER 2026-05-10 · APENAS phones reais PT
                # (mobile 91/92/93/96 ou landline 21-29). Exclui relay/proxy
                # OLX (90X/95X/99X/66X) que dão voicemail anónimo.
                Lead.phone_type.in_(("mobile", "landline")),
            )
            .where(Lead.lead_type != "agency_listing")
            .where(Lead.zone.in_(target_zones))
            .order_by(Lead.score.desc())
            .limit(limit * 5)
        )
        leads = db.execute(q).scalars().all()

    seen_phones: set[str] = set()
    rows: list[dict] = []

    # Sort: adjusted score = real score + phone channel bonus (5 for mobile, 2 for relay).
    # This keeps HOT relay leads above low-score mobile leads, while mobile wins
    # at equal or near-equal scores — matching the user's intent of "mobile first"
    # without demoting high-value relay leads.
    def _sort_key(lead: Lead):
        pt    = getattr(lead, "phone_type", None) or "unknown"
        bonus = 5 if pt == "mobile" else (2 if pt == "relay" else 0)
        return -(( lead.score or 0) + bonus)

    leads_sorted = sorted(leads, key=_sort_key)

    for lead in leads_sorted:
        phone = _canonical_phone(lead.contact_phone or "")
        if not phone:
            continue
        if phone in seen_phones:
            continue
        if _is_excluded_landline(lead):
            continue
        if _is_likely_agency(lead):
            continue
        # Sprint Owner Tier · exclude tier E (switchboard agency confirmed)
        if _classify_owner_tier(lead) == "E":
            continue
        # Filter active_owner + landline — weak signal
        if lead.lead_type == "active_owner" and (
            getattr(lead, "phone_type", None) == "landline"
        ):
            continue

        seen_phones.add(phone)
        row = _lead_to_row(lead, rank=len(rows) + 1)
        rows.append(row)

        if len(rows) >= limit:
            break

    log.info(
        "[commercial] Premium list: {n} leads (from {total} candidates)",
        n=len(rows), total=len(leads),
    )
    return rows


def generate_expanded_list(
    premium_phones: set[str],
    score_min: int | None = None,
    zones: list[str] | None = None,
    limit: int = 150,
) -> list[dict]:
    """
    Build the Lista Expandida.

    Broader criteria than Premium:
      • score ≥ max(score_min, warm_threshold - 10, 25)
      • agencies excluded
      • Lisboa/Porto landlines excluded
      • phones already in Premium excluded
      • dedup by phone
      • sorted by score desc

    Returns list of row dicts, up to `limit` entries.
    """
    from sqlalchemy import select, or_

    warm_t    = settings.warm_score_threshold
    min_score = max(score_min if score_min is not None else warm_t - 10, 25)
    target_zones = zones or settings.zones

    with get_db() as db:
        q = (
            select(Lead)
            .where(
                Lead.archived == False,                       # noqa: E712
                Lead.is_demo  == False,                       # noqa: E712
                Lead.contact_phone.isnot(None),
                Lead.contact_phone != "",
                Lead.score >= min_score,
                (Lead.listing_status.is_(None)) | (Lead.listing_status != "suspicious"),
                # Sprint REAL OWNER · Expandida também só com phones reais
                Lead.phone_type.in_(("mobile", "landline")),
            )
            .where(Lead.lead_type.notin_(("agency_listing",)))
            .where(Lead.zone.in_(target_zones))
            .order_by(Lead.score.desc())
            .limit(limit * 4)
        )
        leads = db.execute(q).scalars().all()

    seen_phones: set[str] = set(premium_phones)   # start from premium exclusions
    rows: list[dict] = []

    for lead in leads:
        phone = _canonical_phone(lead.contact_phone or "")
        if not phone:
            continue
        if phone in seen_phones:
            continue
        if _is_excluded_landline(lead):
            continue
        # Sprint Owner Tier · exclude tier E (switchboard agency confirmed)
        if _classify_owner_tier(lead) == "E":
            continue

        seen_phones.add(phone)
        row = _lead_to_row(lead, rank=len(rows) + 1)
        rows.append(row)

        if len(rows) >= limit:
            break

    log.info(
        "[commercial] Expanded list: {n} leads (from {total} candidates)",
        n=len(rows), total=len(leads),
    )
    return rows


# ── Sprint REAL OWNER · Lista Alternativa ────────────────────────────────────

def generate_alternative_list(
    excluded_phones: set[str],
    score_min: int | None = None,
    zones: list[str] | None = None,
    limit: int = 200,
) -> list[dict]:
    """
    Build the Lista Alternativa.

    Leads with a relay/unknown phone (NOT directly callable) but with a
    valid listing URL — Susana can contact via the portal's internal
    messaging system. Honest disclosure: these are NOT real owner phones.

    Criteria:
      • score ≥ 30 (broader than Premium/Expandida)
      • phone_type IN (relay, unknown) OR no phone
      • has a URL (so Susana can reach via portal)
      • not in Premium/Expandida already
      • zone in target_zones
    """
    from sqlalchemy import select, or_

    target_zones = zones or settings.zones
    min_score    = score_min if score_min is not None else 30

    with get_db() as db:
        q = (
            select(Lead)
            .where(
                Lead.archived == False,                       # noqa: E712
                Lead.is_demo  == False,                       # noqa: E712
                Lead.score >= min_score,
                (Lead.listing_status.is_(None)) | (Lead.listing_status != "suspicious"),
                Lead.zone.in_(target_zones),
                or_(
                    Lead.phone_type.in_(("relay", "unknown", "invalid")),
                    Lead.phone_type.is_(None),
                    Lead.contact_phone.is_(None),
                    Lead.contact_phone == "",
                ),
            )
            .where(Lead.lead_type.notin_(("agency_listing",)))
            .order_by(Lead.score.desc())
            .limit(limit * 3)
        )
        leads = db.execute(q).scalars().all()

    rows: list[dict] = []
    seen: set[str] = set()
    for lead in leads:
        # Skip if phone was already shipped in a previous list
        if lead.contact_phone and lead.contact_phone in excluded_phones:
            continue
        # Need a way to contact — listing URL or email
        url = _get_url(lead.sources_json or "[]")
        if not url and not lead.contact_email:
            continue
        # Cheap dedup by URL
        if url in seen:
            continue
        seen.add(url)

        row = _lead_to_row(lead, rank=len(rows) + 1)
        # Mark explicitly: phone is NOT a real owner number
        ph_t = (lead.phone_type or "—")
        row["telefone"] = (
            "📭 sem telemóvel directo"
            if (not lead.contact_phone) else
            f"⚠ {lead.contact_phone}  (relay/proxy · não chega ao dono)"
        )
        row["mensagem_wa"] = ""    # blank — no real phone to message
        row["whatsapp"]    = ""
        # Highlight URL as the contact channel
        row["url"] = url
        rows.append(row)

        if len(rows) >= limit:
            break

    return rows


# ── Executive summary ─────────────────────────────────────────────────────────

def build_executive_summary(
    premium: list[dict],
    expanded: list[dict],
    generated_at: datetime | None = None,
) -> dict:
    """
    Build a structured executive summary from the two lists.

    Returns a dict suitable for both text rendering and XLSX Sheet 3.
    """
    generated_at = generated_at or datetime.now()
    all_rows = premium + expanded

    # Counts
    p_hot     = sum(1 for r in premium  if r["label"] == "HOT")
    p_warm    = sum(1 for r in premium  if r["label"] == "WARM")
    e_hot     = sum(1 for r in expanded if r["label"] == "HOT")
    e_warm    = sum(1 for r in expanded if r["label"] == "WARM")

    # Phone type breakdown (combined)
    pt_counts: Counter = Counter(r["_phone_type"] for r in all_rows)

    # Lead type breakdown
    lt_counts: Counter = Counter(r["tipo_lead"] for r in all_rows)

    # Top zones (premium only — that's the target)
    zone_counts: Counter = Counter(r["zona"] for r in premium)
    top_zones = zone_counts.most_common(6)

    # Score range in premium
    if premium:
        p_scores = [r["score"] for r in premium]
        p_score_min, p_score_max, p_score_avg = (
            min(p_scores),
            max(p_scores),
            round(sum(p_scores) / len(p_scores), 1),
        )
    else:
        p_score_min = p_score_max = p_score_avg = 0

    return {
        "generated_at":      generated_at.strftime("%d/%m/%Y %H:%M"),
        "total_premium":     len(premium),
        "total_expanded":    len(expanded),
        "total_combined":    len(all_rows),
        # Premium breakdown
        "premium_hot":       p_hot,
        "premium_warm":      p_warm,
        "premium_cold":      len(premium) - p_hot - p_warm,
        "premium_score_min": p_score_min,
        "premium_score_max": p_score_max,
        "premium_score_avg": p_score_avg,
        # Expanded breakdown
        "expanded_hot":      e_hot,
        "expanded_warm":     e_warm,
        "expanded_cold":     len(expanded) - e_hot - e_warm,
        # Phone channels
        "mobile_count":      pt_counts.get("mobile", 0),
        "relay_count":       pt_counts.get("relay", 0),
        "landline_count":    pt_counts.get("landline", 0),
        "unknown_count":     pt_counts.get("unknown", 0),
        # Lead types
        "lead_type_counts":  dict(lt_counts.most_common()),
        # Zones
        "top_zones":         top_zones,
    }


def summary_as_text(summary: dict) -> str:
    """Render the executive summary as a human-readable text block."""
    sep  = "─" * 58
    sep2 = "═" * 58
    top_zones_str = "  ".join(
        f"{z} ({n})" for z, n in summary["top_zones"]
    )

    lt = summary["lead_type_counts"]
    lt_str = "  ".join(f"{k}: {v}" for k, v in lt.items())

    lines = [
        "",
        sep2,
        "  RELATÓRIO COMERCIAL — LISTA DE LEADS",
        f"  Gerado em: {summary['generated_at']}",
        sep2,
        "",
        f"  LISTA PREMIUM ({summary['total_premium']} contactos únicos)",
        sep,
        f"  🔴 HOT          : {summary['premium_hot']}",
        f"  🟡 WARM         : {summary['premium_warm']}",
        f"  Score (min/avg/max) : {summary['premium_score_min']} / "
        f"{summary['premium_score_avg']} / {summary['premium_score_max']}",
        "",
        f"  LISTA EXPANDIDA ({summary['total_expanded']} contactos únicos)",
        sep,
        f"  🔴 HOT          : {summary['expanded_hot']}",
        f"  🟡 WARM         : {summary['expanded_warm']}",
        "",
        "  CANAIS DE CONTACTO (combinado)",
        sep,
        f"  📱 Telemóvel    : {summary['mobile_count']}",
        f"  🔁 Relay/OLX    : {summary['relay_count']}",
        f"  📞 Fixo         : {summary['landline_count']}",
        f"  ❓ Desconhecido : {summary['unknown_count']}",
        "",
        "  TIPOS DE LEAD",
        sep,
        f"  {lt_str}",
        "",
        "  PRINCIPAIS ZONAS (Premium)",
        sep,
        f"  {top_zones_str}",
        "",
        sep2,
    ]
    return "\n".join(lines)


# ── XLSX export ───────────────────────────────────────────────────────────────

def export_commercial_xlsx(
    premium:  list[dict],
    expanded: list[dict],
    summary:  dict,
    output_path: str,
    alternative: list[dict] | None = None,
) -> str:
    """
    Generate a single client-ready XLSX with three sheets.

    Sheet 1: Lista Premium
    Sheet 2: Lista Expandida
    Sheet 3: Resumo Executivo
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")

    wb = openpyxl.Workbook()

    # ── Colour palette ────────────────────────────────────────────────────────
    C = {
        "header_bg":    "1F3864",
        "header_fg":    "FFFFFF",
        "hot_bg":       "FFD6D6",
        "warm_bg":      "FFF4CC",
        "mobile_bg":    "E8F5E9",   # soft green — mobile numbers
        "relay_bg":     "E3F2FD",   # soft blue — relay numbers
        "alt_bg":       "F8FBFF",
        "white":        "FFFFFF",
        "border":       "D0D7E0",
        "premium_tab":  "C62828",
        "expanded_tab": "1565C0",
        "summary_tab":  "1B5E20",
    }

    thin = Border(
        left=Side(style="thin",  color=C["border"]),
        right=Side(style="thin", color=C["border"]),
        top=Side(style="thin",   color=C["border"]),
        bottom=Side(style="thin",color=C["border"]),
    )

    # ── Column definitions (shared between Premium and Expandida) ─────────────
    COLUMNS = [
        ("#",            "rank",          4),
        ("Score",        "score",         7),
        ("Label",        "label",         8),
        ("Tier",         "owner_tier",    6),
        ("Sinais",       "badges",       30),
        ("Perfil Vendedor","owner_profile",30),
        ("Nome",         "nome",         22),
        ("Telefone",     "telefone",     16),
        ("Tipo Tel.",    "tipo_telefone",12),
        ("WhatsApp",     "whatsapp",     18),
        ("Mensagem WA",  "mensagem_wa",  60),
        ("Zona",         "zona",         12),
        ("Tipologia",    "tipologia",    10),
        ("Preço",        "preco",        14),
        ("Área",         "area_m2",       9),
        ("Tipo Lead",    "tipo_lead",    22),
        ("Fonte",        "fonte",        14),
        ("Confiança",    "confianca",    10),
        ("Insight",      "insight",      42),
        ("URL Anúncio",  "url",          18),
        ("Dias Merc.",   "dias_mercado",  9),
        ("Captado em",   "data_captacao",12),
        ("Comissão est.","comissao_est", 14),
        ("Opening",      "opening",      60),
        ("Briefing",     "briefing",     80),
    ]

    def _write_sheet(ws, rows: list[dict], title: str, tab_color: str) -> None:
        ws.title = title
        ws.sheet_properties.tabColor = tab_color

        # Header
        h_font  = Font(name="Calibri", bold=True, color=C["header_fg"], size=10)
        h_fill  = PatternFill("solid", fgColor=C["header_bg"])
        h_align = Alignment(horizontal="center", vertical="center")

        for ci, (hdr, _, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font  = h_font
            cell.fill  = h_fill
            cell.border = thin
            cell.alignment = h_align
            ws.column_dimensions[get_column_letter(ci)].width = width

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        # Data rows
        d_font    = Font(name="Calibri", size=9)
        bold_font = Font(name="Calibri", size=9, bold=True)
        link_font = Font(name="Calibri", size=9, color="1565C0", underline="single")
        c_align   = Alignment(horizontal="center", vertical="center")
        l_align   = Alignment(horizontal="left",   vertical="center")
        r_align   = Alignment(horizontal="right",  vertical="center")
        wrap_al   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        for ri, row in enumerate(rows, 2):
            pt    = row.get("_phone_type", "unknown")
            label = row.get("label", "COLD")

            # Row background: HOT > WARM > mobile > relay > alternating
            if label == "HOT":
                bg = C["hot_bg"]
            elif label == "WARM":
                bg = C["warm_bg"]
            elif pt == "mobile":
                bg = C["mobile_bg"]
            elif pt == "relay":
                bg = C["relay_bg"]
            elif ri % 2 == 0:
                bg = C["alt_bg"]
            else:
                bg = C["white"]

            row_fill = PatternFill("solid", fgColor=bg)

            for ci, (_, key, _) in enumerate(COLUMNS, 1):
                v    = row.get(key, "")
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = thin
                cell.fill   = row_fill

                if key == "rank":
                    cell.font = Font(name="Calibri", size=9, color="888888")
                    cell.alignment = c_align
                elif key == "score":
                    cell.font      = bold_font
                    cell.alignment = r_align
                elif key == "label":
                    color = {"HOT": "C00000", "WARM": "B8860B"}.get(label, "444444")
                    cell.font      = Font(name="Calibri", size=9, bold=True, color=color)
                    cell.alignment = c_align
                elif key == "tipo_telefone":
                    icon = {"Telemóvel": "📱", "Relay/OLX": "🔁",
                            "Fixo": "📞", "Desconhecido": "❓"}.get(str(v), "")
                    cell.value     = f"{icon} {v}" if icon else v
                    cell.font      = d_font
                    cell.alignment = c_align
                elif key == "whatsapp":
                    if v and v.startswith("http"):
                        cell.hyperlink = v
                        cell.value     = "📲 WhatsApp"
                        cell.font      = link_font
                    cell.alignment = c_align
                elif key == "url":
                    if v and v.startswith("http"):
                        cell.hyperlink = v
                        cell.value     = "🔗 Ver Anúncio"
                        cell.font      = link_font
                    cell.alignment = c_align
                elif key == "insight":
                    cell.font      = Font(name="Calibri", size=9, italic=True)
                    cell.alignment = wrap_al
                elif key == "confianca":
                    color = {"ALTA": "1B5E20", "MÉDIA": "E65100", "BAIXA": "B71C1C"}.get(str(v), "444444")
                    cell.font      = Font(name="Calibri", size=9, bold=True, color=color)
                    cell.alignment = c_align
                elif key == "dias_mercado":
                    cell.font      = d_font
                    cell.alignment = r_align
                elif key in ("briefing", "opening"):
                    cell.font      = d_font
                    cell.alignment = wrap_al
                else:
                    cell.font      = d_font
                    cell.alignment = l_align

            # Tier A/B leads have a multi-line briefing — taller row so it shows
            ws.row_dimensions[ri].height = 110 if row.get("briefing") and row["briefing"] != "—" else 18

        # Auto-filter on header row
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Sheet 1: Premium ──────────────────────────────────────────────────────
    ws_premium = wb.active
    _write_sheet(ws_premium, premium, "Lista Premium", C["premium_tab"])

    # ── Sheet 2: Expandida ────────────────────────────────────────────────────
    ws_expanded = wb.create_sheet()
    _write_sheet(ws_expanded, expanded, "Lista Expandida", C["expanded_tab"])

    # ── Sheet 3: ⚠ SEM TELEFONE (Sprint REAL OWNER) ──────────────────────────
    # Leads SEM phone real (relay/proxy/none), contactáveis via portal URL.
    # Susana liga via mensagem interna do OLX/Imovirtual em vez de telefone.
    if alternative:
        ws_alt = wb.create_sheet()
        _write_sheet(ws_alt, alternative, "⚠ SEM TELEFONE", "F0AD4E")  # warning amber
        # Insert a banner row at top explaining the sheet
        ws_alt.insert_rows(1, amount=2)
        banner = ws_alt.cell(row=1, column=1,
            value="⚠  ATENÇÃO: ESTES LEADS NÃO TÊM TELEFONE REAL")
        banner.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        banner.fill = PatternFill("solid", fgColor="C62828")
        banner.alignment = Alignment(horizontal="center", vertical="center")
        ws_alt.merge_cells(start_row=1, start_column=1,
                            end_row=1, end_column=len(COLUMNS))
        ws_alt.row_dimensions[1].height = 28

        sub = ws_alt.cell(row=2, column=1,
            value="O número mostrado é um relay/proxy do portal — NÃO é o número do dono. "
                  "Para contactar: clica no URL do anúncio e envia mensagem interna no OLX/Imovirtual. "
                  "NÃO ligar — o relay não atende ou redirecciona para sistema automático.")
        sub.font = Font(name="Calibri", italic=True, size=10, color="C62828")
        sub.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws_alt.merge_cells(start_row=2, start_column=1,
                            end_row=2, end_column=len(COLUMNS))
        ws_alt.row_dimensions[2].height = 36
        # Move freeze pane down so banner stays visible
        ws_alt.freeze_panes = "A4"

    # ── Sheet 4: Resumo Executivo ─────────────────────────────────────────────
    ws_sum = wb.create_sheet("Resumo Executivo")
    ws_sum.sheet_properties.tabColor = C["summary_tab"]

    title_font  = Font(name="Calibri", bold=True, size=14, color=C["header_bg"])
    section_font= Font(name="Calibri", bold=True, size=10, color="1F3864")
    label_font  = Font(name="Calibri", size=10)
    value_font  = Font(name="Calibri", bold=True, size=10)
    note_font   = Font(name="Calibri", size=9, italic=True, color="666666")

    ws_sum.column_dimensions["A"].width = 32
    ws_sum.column_dimensions["B"].width = 20
    ws_sum.column_dimensions["C"].width = 28

    def _sum_row(r, col_a, col_b="", col_c="", bold=False, section=False, note=False):
        ca = ws_sum.cell(row=r, column=1, value=col_a)
        cb = ws_sum.cell(row=r, column=2, value=col_b)
        cc = ws_sum.cell(row=r, column=3, value=col_c)
        if section:
            ca.font = section_font
            ca.fill = PatternFill("solid", fgColor="EEF2F8")
        elif bold:
            ca.font = value_font
            cb.font = value_font
        elif note:
            ca.font = note_font
            cb.font = note_font
        else:
            ca.font = label_font
            cb.font = value_font
            cc.font = label_font
        return r + 1

    r = 1
    ws_sum.cell(row=r, column=1, value="Relatório Comercial de Leads").font = title_font
    r += 1
    ws_sum.cell(row=r, column=1, value=f"Gerado em: {summary['generated_at']}").font = note_font
    r += 2

    r = _sum_row(r, "LISTA PREMIUM", section=True)
    r = _sum_row(r, "Total contactos únicos",   summary["total_premium"])
    r = _sum_row(r, "  🔴 HOT (score ≥ 60)",   summary["premium_hot"])
    r = _sum_row(r, "  🟡 WARM (score 40-59)",  summary["premium_warm"])
    r = _sum_row(r, "  Score min / avg / max",
                 f"{summary['premium_score_min']} / {summary['premium_score_avg']} / {summary['premium_score_max']}")
    r += 1

    r = _sum_row(r, "LISTA EXPANDIDA", section=True)
    r = _sum_row(r, "Total contactos únicos",   summary["total_expanded"])
    r = _sum_row(r, "  🔴 HOT",                 summary["expanded_hot"])
    r = _sum_row(r, "  🟡 WARM",                summary["expanded_warm"])
    r += 1

    r = _sum_row(r, "CANAIS DE CONTACTO (combinado)", section=True)
    r = _sum_row(r, "📱 Telemóvel",    summary["mobile_count"])
    r = _sum_row(r, "🔁 Relay/OLX",   summary["relay_count"])
    r = _sum_row(r, "📞 Fixo",         summary["landline_count"])
    r = _sum_row(r, "❓ Desconhecido", summary["unknown_count"])
    r += 1

    r = _sum_row(r, "TIPOS DE LEAD", section=True)
    for lt_label, lt_count in summary["lead_type_counts"].items():
        r = _sum_row(r, f"  {lt_label}", lt_count)
    r += 1

    r = _sum_row(r, "PRINCIPAIS ZONAS (Premium)", section=True)
    for zone, count in summary["top_zones"]:
        bar = "█" * min(count, 20)
        r = _sum_row(r, f"  {zone}", count, bar)
    r += 1

    r = _sum_row(r, "TOTAL GERAL", summary["total_combined"], bold=True)

    r += 2
    _sum_row(r, "Legenda cores Premium/Expandida:", note=True)
    r += 1
    _sum_row(r, "  Vermelho claro = HOT (score ≥ 60)", note=True)
    r += 1
    _sum_row(r, "  Amarelo claro = WARM (score 40–59)", note=True)
    r += 1
    _sum_row(r, "  Verde claro = Telemóvel directo", note=True)
    r += 1
    _sum_row(r, "  Azul claro = Relay/OLX", note=True)

    wb.save(output_path)
    log.info("[commercial] XLSX saved → {p}", p=output_path)
    return output_path


# ── CSV export (fallback / two separate files) ────────────────────────────────

def export_commercial_csv(
    rows: list[dict],
    output_path: str,
) -> str:
    """Export a single list to CSV (utf-8-sig for Excel compatibility)."""
    import csv
    fields = [
        "rank", "score", "label", "nome", "telefone", "tipo_telefone",
        "whatsapp", "zona", "concelho", "tipologia", "preco", "area_m2",
        "tipo_lead", "fonte", "confianca", "insight", "url",
        "dias_mercado", "data_captacao",
    ]
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    log.info("[commercial] CSV saved → {p} ({n} rows)", p=output_path, n=len(rows))
    return output_path


# ── Main orchestrator ─────────────────────────────────────────────────────────

def run_commercial_export(
    premium_limit:  int = 50,
    expanded_limit: int = 150,
    zones: list[str] | None = None,
    fmt: str = "xlsx",           # "xlsx" | "csv" | "both"
    output_dir: str | None = None,
) -> dict:
    """
    Full commercial export flow.

    1. Build Lista Premium
    2. Build Lista Expandida (excluding Premium phones)
    3. Build Executive Summary
    4. Export to XLSX (single file, 3 sheets) and/or CSV (2 files)

    Returns:
        {
            "premium":  list[dict],
            "expanded": list[dict],
            "summary":  dict,
            "files":    {"xlsx": str, "csv_premium": str, "csv_expanded": str},
        }
    """
    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = Path(output_dir) if output_dir else Path(settings.data_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    log.info("[commercial] Building Premium list (limit={n})…", n=premium_limit)
    premium = generate_premium_list(zones=zones, limit=premium_limit)

    premium_phones = {r["telefone"] for r in premium if r.get("telefone")}

    log.info("[commercial] Building Expanded list (limit={n})…", n=expanded_limit)
    expanded = generate_expanded_list(
        premium_phones=premium_phones,
        zones=zones,
        limit=expanded_limit,
    )

    # Sprint REAL OWNER · Lista Alternativa (sem phone real, contactáveis via portal)
    expanded_phones = {r["telefone"] for r in expanded if r.get("telefone")}
    log.info("[commercial] Building Lista Alternativa (relay/no-phone leads)…")
    alternative = generate_alternative_list(
        excluded_phones=premium_phones | expanded_phones,
        zones=zones,
        limit=200,
    )
    log.info("[commercial] Lista Alternativa: {n} leads", n=len(alternative))

    summary = build_executive_summary(premium, expanded)
    summary["alternative_count"] = len(alternative)

    log.info(
        "[commercial] Summary — Premium: {p} | Expanded: {e} | "
        "Mobile: {m} | Relay: {r}",
        p=len(premium), e=len(expanded),
        m=summary["mobile_count"], r=summary["relay_count"],
    )

    files: dict[str, str] = {}

    if fmt in ("xlsx", "both"):
        xlsx_path = str(out_dir / f"leads_comercial_{ts}.xlsx")
        try:
            export_commercial_xlsx(premium, expanded, summary, xlsx_path, alternative=alternative)
            files["xlsx"] = xlsx_path
        except ImportError as exc:
            log.warning("XLSX export skipped: {e}", e=exc)

    if fmt in ("csv", "both"):
        csv_p = str(out_dir / f"leads_premium_{ts}.csv")
        csv_e = str(out_dir / f"leads_expandida_{ts}.csv")
        export_commercial_csv(premium,  csv_p)
        export_commercial_csv(expanded, csv_e)
        files["csv_premium"]  = csv_p
        files["csv_expanded"] = csv_e

    return {
        "premium":  premium,
        "expanded": expanded,
        "summary":  summary,
        "files":    files,
    }
