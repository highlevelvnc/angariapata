"""
Conversion Analytics · Sprint 2026-05-11

Closes the feedback loop by surfacing conversion KPIs from the data Susana
fills back into the system. Answers questions like:

  - Que % dos Tier A respondem? E convertem?
  - Que zona dá mais visitas marcadas?
  - O OLX converte melhor que Imovirtual?
  - Quantos dias entre 1ª chamada e visita marcada?
  - Que faixa de preço converte mais?

Output:
  - Rich console table (resumo executivo)
  - XLSX opcional com 6 folhas de breakdown (--out path.xlsx)

Pure offline aggregation — só lê `crm_stage`, `last_contacted_at`,
`contact_outcome` da DB. Não dispara nenhuma chamada externa.

Conversion stages (em ordem de valor crescente):
  novo            → ainda não contactado (não conta)
  sem_resposta    → não atendeu (no_answer)
  contactado      → atendeu mas sem decisão (reached)
  indisponivel    → número morto (dead)
  nao_interessado → recusou (refused)
  interessado     → mostrou interesse (warm_outcome) 🟡
  convertido      → angariação fechada (won) ✅

A "conversion" = stage IN (interessado, convertido).
"""
from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

from sqlalchemy import func, or_

from storage.database import get_db
from storage.models   import Lead
from utils.logger     import get_logger
from reports.commercial_export import _classify_owner_tier, _short_zone

log = get_logger(__name__)


# ── Stage taxonomy ───────────────────────────────────────────────────────────

_OUTCOME_WEIGHTS = {
    "convertido":      "won",
    "interessado":     "warm_outcome",
    "contactado":      "reached",
    "sem_resposta":    "no_answer",
    "indisponivel":    "dead",
    "nao_interessado": "refused",
    "novo":            "untouched",
}

_CONVERSION_STAGES = {"interessado", "convertido"}
_CONTACT_ATTEMPTED = {
    "sem_resposta", "contactado", "indisponivel",
    "nao_interessado", "interessado", "convertido",
}


@dataclass
class Bucket:
    """One row of analytics — a bucket (zone / source / tier / etc.)."""
    label:        str
    total_leads:  int = 0
    contacted:    int = 0
    converted:    int = 0
    refused:      int = 0
    no_answer:    int = 0
    dead:         int = 0

    @property
    def contact_rate(self) -> float:
        return self.contacted / self.total_leads if self.total_leads else 0.0

    @property
    def conversion_rate(self) -> float:
        return self.converted / self.contacted if self.contacted else 0.0

    @property
    def conversion_pct(self) -> str:
        return f"{self.conversion_rate * 100:.1f}%" if self.contacted else "—"


# ── Aggregation ──────────────────────────────────────────────────────────────

def _bucket_lead(b: Bucket, lead: Lead) -> None:
    b.total_leads += 1
    stage = (lead.crm_stage or "novo").lower()
    if stage in _CONTACT_ATTEMPTED:
        b.contacted += 1
    if stage in _CONVERSION_STAGES:
        b.converted += 1
    if stage == "nao_interessado":
        b.refused += 1
    if stage == "sem_resposta":
        b.no_answer += 1
    if stage == "indisponivel":
        b.dead += 1


def compute_analytics(
    score_min: int = 0,
    zones: Optional[list[str]] = None,
) -> dict:
    """
    Aggregate conversion metrics across all non-archived leads (and
    archived leads that were touched — so converted/refused stay visible).

    Returns a dict with sub-bucket dicts: by_zone, by_source, by_tier,
    by_score_band, by_phone_type, by_lead_type, plus an `overall` Bucket.
    """
    by_zone:       dict[str, Bucket] = defaultdict(lambda: Bucket(""))
    by_source:     dict[str, Bucket] = defaultdict(lambda: Bucket(""))
    by_tier:       dict[str, Bucket] = defaultdict(lambda: Bucket(""))
    by_score_band: dict[str, Bucket] = defaultdict(lambda: Bucket(""))
    by_phone:      dict[str, Bucket] = defaultdict(lambda: Bucket(""))
    by_lead_type:  dict[str, Bucket] = defaultdict(lambda: Bucket(""))
    overall = Bucket("overall")
    stage_counts: Counter[str] = Counter()
    days_to_conv: list[int] = []   # days between first_seen_at → last_contacted_at for converted

    with get_db() as db:
        # Include archived too — we want to count converted/refused
        q = db.query(Lead).filter(Lead.score >= score_min)
        if zones:
            q = q.filter(Lead.zone.in_(zones))
        leads = q.all()

        for L in leads:
            stage = (L.crm_stage or "novo").lower()
            stage_counts[stage] += 1

            _bucket_lead(overall, L)

            z = _short_zone(L.zone) or "—"
            if not by_zone[z].label:
                by_zone[z].label = z
            _bucket_lead(by_zone[z], L)

            src = (L.discovery_source or "—").lower()
            if not by_source[src].label:
                by_source[src].label = src
            _bucket_lead(by_source[src], L)

            tier = _classify_owner_tier(L) if L.contact_phone else "?"
            if not by_tier[tier].label:
                by_tier[tier].label = tier
            _bucket_lead(by_tier[tier], L)

            s = L.score or 0
            band = ("HOT" if s >= 60 else "WARM" if s >= 40
                    else "COLD" if s >= 20 else "FRIO")
            if not by_score_band[band].label:
                by_score_band[band].label = band
            _bucket_lead(by_score_band[band], L)

            pt = (L.phone_type or "unknown").lower()
            if not by_phone[pt].label:
                by_phone[pt].label = pt
            _bucket_lead(by_phone[pt], L)

            lt = (L.lead_type or "unknown").lower()
            if not by_lead_type[lt].label:
                by_lead_type[lt].label = lt
            _bucket_lead(by_lead_type[lt], L)

            # Time-to-conversion (days)
            if stage in _CONVERSION_STAGES and L.last_contacted_at and L.first_seen_at:
                delta = (L.last_contacted_at - L.first_seen_at).days
                if delta >= 0:
                    days_to_conv.append(delta)

    avg_days = (sum(days_to_conv) / len(days_to_conv)) if days_to_conv else 0.0

    return {
        "overall":       overall,
        "stage_counts":  dict(stage_counts),
        "by_zone":       dict(by_zone),
        "by_source":     dict(by_source),
        "by_tier":       dict(by_tier),
        "by_score_band": dict(by_score_band),
        "by_phone_type": dict(by_phone),
        "by_lead_type":  dict(by_lead_type),
        "days_to_conversion_avg":   avg_days,
        "days_to_conversion_count": len(days_to_conv),
        "generated_at": datetime.utcnow(),
    }


# ── Renderers ────────────────────────────────────────────────────────────────

def render_console(stats: dict) -> None:
    """Pretty-print to the terminal using rich."""
    from rich.console import Console
    from rich.table   import Table
    from rich         import box

    console = Console()
    overall: Bucket = stats["overall"]

    console.print()
    console.print("[bold cyan]═══ CONVERSION ANALYTICS ═══[/bold cyan]")
    console.print(
        f"Total leads : [bold]{overall.total_leads}[/bold]   "
        f"Contactados : [bold]{overall.contacted}[/bold]   "
        f"Convertidos : [bold green]{overall.converted}[/bold green]   "
        f"Recusados : [bold red]{overall.refused}[/bold red]"
    )
    if overall.contacted:
        console.print(
            f"Contact rate: {overall.contact_rate*100:.1f}%   "
            f"Conversion rate: [bold green]{overall.conversion_rate*100:.1f}%[/bold green]"
        )
    if stats["days_to_conversion_count"]:
        console.print(
            f"Tempo médio até conversão: [bold]{stats['days_to_conversion_avg']:.1f}d[/bold] "
            f"({stats['days_to_conversion_count']} conversões medidas)"
        )

    # Stage breakdown
    sc = stats["stage_counts"]
    stage_tbl = Table(title="Estados", box=box.SIMPLE_HEAD)
    stage_tbl.add_column("Estado")
    stage_tbl.add_column("Total", justify="right")
    for stage in ("convertido", "interessado", "contactado", "sem_resposta",
                  "indisponivel", "nao_interessado", "novo"):
        if sc.get(stage):
            stage_tbl.add_row(stage, str(sc[stage]))
    console.print(stage_tbl)

    # Helper renderer
    def _render(title: str, buckets: dict[str, Bucket], top: int = 12):
        if not buckets:
            return
        rows = sorted(buckets.values(),
                      key=lambda b: (b.contacted, b.converted), reverse=True)
        rows = [b for b in rows if b.total_leads >= 5][:top]
        if not rows:
            return
        tbl = Table(title=title, box=box.SIMPLE_HEAD)
        tbl.add_column(title.replace("Por ", ""))
        tbl.add_column("Total", justify="right")
        tbl.add_column("Contactados", justify="right")
        tbl.add_column("Convertidos", justify="right")
        tbl.add_column("Conv %", justify="right")
        for b in rows:
            tbl.add_row(b.label, str(b.total_leads), str(b.contacted),
                        str(b.converted), b.conversion_pct)
        console.print(tbl)

    _render("Por Zona",        stats["by_zone"])
    _render("Por Source",      stats["by_source"])
    _render("Por Tier",        stats["by_tier"])
    _render("Por Score Band",  stats["by_score_band"])
    _render("Por Phone Type",  stats["by_phone_type"])
    _render("Por Lead Type",   stats["by_lead_type"])


def export_xlsx(stats: dict, output_path: str) -> str:
    """Write a 6-sheet XLSX with the breakdowns."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    overall: Bucket = stats["overall"]

    def _write_table(ws_name: str, buckets: dict[str, Bucket], col_label: str):
        ws = wb.create_sheet(ws_name)
        headers = [col_label, "Total Leads", "Contactados", "Convertidos",
                   "Recusados", "Sem Resposta", "Mortos",
                   "Contact %", "Conversion %"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F3864")
            c.alignment = Alignment(horizontal="center")
        rows = sorted(buckets.values(),
                      key=lambda b: b.total_leads, reverse=True)
        for r, b in enumerate(rows, 2):
            ws.cell(row=r, column=1, value=b.label)
            ws.cell(row=r, column=2, value=b.total_leads)
            ws.cell(row=r, column=3, value=b.contacted)
            ws.cell(row=r, column=4, value=b.converted)
            ws.cell(row=r, column=5, value=b.refused)
            ws.cell(row=r, column=6, value=b.no_answer)
            ws.cell(row=r, column=7, value=b.dead)
            ws.cell(row=r, column=8, value=f"{b.contact_rate*100:.1f}%")
            ws.cell(row=r, column=9, value=b.conversion_pct)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 18

    # Overview sheet first
    ws0 = wb.active
    ws0.title = "Resumo"
    ws0["A1"] = "CONVERSION ANALYTICS — Pata Brava"
    ws0["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws0["A3"] = "Gerado em"
    ws0["B3"] = stats["generated_at"].strftime("%d/%m/%Y %H:%M")
    metrics = [
        ("Total leads (todos)",          overall.total_leads),
        ("Contactados",                  overall.contacted),
        ("Convertidos (interessado + ✓)", overall.converted),
        ("Recusados",                    overall.refused),
        ("Sem resposta",                 overall.no_answer),
        ("Números mortos",               overall.dead),
        ("Contact rate %",               f"{overall.contact_rate*100:.1f}%"),
        ("Conversion rate %",            overall.conversion_pct),
        ("Tempo médio até conversão (d)", f"{stats['days_to_conversion_avg']:.1f}"),
        ("# conversões medidas",         stats["days_to_conversion_count"]),
    ]
    for i, (label, val) in enumerate(metrics, 5):
        ws0.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws0.cell(row=i, column=2, value=val)
    ws0.column_dimensions["A"].width = 36
    ws0.column_dimensions["B"].width = 18

    _write_table("Por Zona",       stats["by_zone"],       "Zona")
    _write_table("Por Source",     stats["by_source"],     "Source")
    _write_table("Por Tier",       stats["by_tier"],       "Tier")
    _write_table("Por Score Band", stats["by_score_band"], "Score Band")
    _write_table("Por Phone Type", stats["by_phone_type"], "Phone Type")
    _write_table("Por Lead Type",  stats["by_lead_type"],  "Lead Type")

    wb.save(output_path)
    log.info("[conversion_analytics] wrote {p}", p=output_path)
    return output_path
